import openpyxl

from openpyxl import Workbook, load_workbook 
from openpyxl.utils import get_column_letter

teamArray = [253, 670, 840, 1072, 1148, 1323, 1422, 1671, 1828, 2135, 2288, 2473, 2551, 3189, 3970, 4400, 4643, 4738, 5026, 5104, 5817, 5940, 6238, 6305, 6657, 6662, 6711, 6884, 6918, 7419, 7777, 8033, 8048, 8262, 8751, 8768, 9202, 9274]
keyList = ["TN", "MN", "AC", "DSN", "SN", "ACOH", "ACOM", "ACOL", "ACUH", "ACUM", "ACUL", "AM", "ACS", "TCOH", "TCOM", "TCOL", "TCUH", "TCUM", "TCUL", "TCS", "DE", "TM", "TS", "DI", "B", "C", "WL", "RPS"]
arrar = [100, 2, "Red", 1, "Kash", 9, 4, 1, 7, 3, 7, 10, "Docked", 9, 8, 7, 5, 3, 9, "Engaged", 4, 6, 4, 7, 1, 3, "win", 2]

def addData(dict, wb, listKey): 

    ws = wb[str(dict["TN"])]

    colWanted = addMatch2(ws, str(dict["MN"]))

    for i in range(0, len(listKey) - 2 ):
        
        rowWanted = getRow(ws, 1, listKey[i + 2])
        
        ws[str(colWanted) + str(rowWanted)].value = dict[listKey[i + 2]]

    totalConeRow = getRow(ws, 1, "TCOS")
    totalCubeRow = getRow(ws, 1, "TCUS")
    
    totalConeNum = dict["TCOH"] + dict["TCOM"] + dict["TCOL"]
    totalCubeNum = dict["TCUH"] + dict["TCUM"] + dict["TCUL"]


    ws[str(colWanted) + str(totalConeRow)].value = totalConeNum
    ws[str(colWanted) + str(totalCubeRow)].value = totalCubeNum


    # Full Data
    fullData(wb, dict)

def fullData(wb, dict):
    fullDataWS = wb["Full Data"]
    fullRowWanted = getRow(fullDataWS, 1, dict["TN"])

    # count
    countCol = getCol(fullDataWS, 1, "CNT")

    # print(get_column_letter(countCol))
    # print(str(fullRowWanted))
    pastCountVal = fullDataWS[ str(get_column_letter(countCol)) + str(fullRowWanted)].value
    
    newCountVal = pastCountVal + 1
    fullDataWS[ str(get_column_letter(countCol)) + str(fullRowWanted) ].value = newCountVal

    for i in range(2, 11):
        char = get_column_letter(i)
        ticker = str(fullDataWS[str(char) + str(1)].value)
        
        pastAvg = fullDataWS[str(char) + str(fullRowWanted)].value
        
       
        if(ticker == "AVCUH"):
            newVal = int(dict["ACUH"]) + int(dict["TCUH"])
        
        if(ticker == "AVCUM"):
            newVal = int(dict["ACUM"]) + int(dict["TCUM"])
        
        if(ticker == "AVCUL"):
            newVal = int(dict["ACUL"]) + int(dict["TCUL"])

        if(ticker == "AVCOH"):
            newVal = int(dict["ACOH"]) + int(dict["TCOH"])

        if(ticker == "AVCOM"):
            newVal = int(dict["ACOM"]) + int(dict["TCOM"])
        
        if(ticker == "AVCOL"):
            
            newVal = int(dict["ACOL"]) + int(dict["TCOL"])

        if(ticker == "ARPS"):
            newVal = int(dict["RPS"])

        if(ticker == "WLP"):
            if(dict["WL"] == "win"):
                newVal = 100
            else:
                newVal = 0

        if(ticker == "CS"):
            if(dict["ACS"] == "Docked"):
                autoCS = 8
            elif(dict["ACS"] == "Engaged"):
                autoCS = 12

            if(dict["TCS"] == "Docked"):
                teleopCS = 6
            elif(dict["TCS"] == "Engaged"):
                teleopCS = 10
            
            newVal = autoCS + teleopCS
       
        

        pastVal = pastAvg * pastCountVal
        newComb = int(newVal) + int(pastVal)
        newAvg = newComb / newCountVal
    
        fullDataWS[str(char) + str(fullRowWanted)].value = newAvg


def getCol(ws, row, string):
    for col in range(1, 100):
        char = get_column_letter(col)
        if(ws[char + str(row)].value == string):
            return col
    return -1


def getRow(ws, col, string):

    for row in range(1, 100):
        char = get_column_letter(col)
        if(ws[char + str(row)].value == string):
            return row
    return -1

def addMatch2(ws, matchNumber):
    newCol = ""
    for col in range(1, 100):

        char = get_column_letter(col)

        if(ws[char + str(1)].value is None):
            newCol = char
            break

    ws[newCol + "1"].value = "Match " + str(matchNumber)
    return newCol

def getRangeRows(ws, col):
    newRow = ""
    char = get_column_letter(col)
    for row in range(1, 1000):
        if(ws[char + str(row)].value is None):
            newRow = row
            break
    return newRow

def getRangeCol(ws, row):
    newCol = ""
    
    for col in range(1, 1000):
        char =  get_column_letter(col)
        
        if(ws[char + str(row)].value is None):
            newCol = char
            break
    return newCol

def makeSheet(wb, arr):

    for i in range(0, len(arr)):
        new_sheet = wb.create_sheet(str(arr[i]))

def fillData(dict, arr, listKey):
    newDict = dict
    for i in range(0, len(dict)):
        newDict[listKey[i]] = arr[i]
    return newDict

def arrToDict(arr, listKey):
    emptyRecord = dict(zip(listKey, [None]*len(listKey)))
    dataRecord = fillData(emptyRecord, arr, listKey)
    return dataRecord

def arrToData(arr, listKey):
    dataRecord = arrToDict(arr, listKey)
    print(dataRecord)
    addData(dataRecord, dataBook, listKey)



# actual code starts here
  
dataBook = load_workbook("database_v2.xlsx")

arrToData(arrar, keyList)

dataBook.save("database_v2.xlsx")



